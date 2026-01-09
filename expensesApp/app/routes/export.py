"""
Export Routes
-------------
Export expenses to Excel format compatible with Google Sheets.
"""
from datetime import date, datetime
from io import BytesIO
from flask import Blueprint, send_file, request
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..models.expense import FoodExpense, UtilityExpense, StuffExpense, OtherExpense
from ..models.fixed_expense import FixedExpense
from ..config import get_config


bp = Blueprint('export', __name__, url_prefix='/export')


@bp.route('/excel')
@login_required
def export_excel():
    """
    Export all expenses to Excel file.
    Structured for Google Sheets compatibility.
    """
    config = get_config()
    selected_year = request.args.get('year', date.today().year, type=int)
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    money_format = '€#,##0.00'
    date_format = 'DD/MM/YYYY'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === Sheet 1: Monthly Summary ===
    ws_summary = wb.active
    ws_summary.title = "Monthly Summary"
    
    # Headers
    headers = ['Month', 'Rent', 'Utilities', 'Food', 'Stuff', 'Other', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Data rows
    month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']
    
    for month in range(1, 13):
        row = month + 1
        rent = FixedExpense.get_value_for_month('Rent', selected_year, month)
        internet = FixedExpense.get_value_for_month('Internet', selected_year, month)
        utilities = UtilityExpense.get_total_by_month(selected_year, month) + internet
        food = FoodExpense.get_total_by_month(selected_year, month)
        stuff = StuffExpense.get_total_by_month(selected_year, month)
        other = OtherExpense.get_total_by_month(selected_year, month)
        total = rent + utilities + food + stuff + other
        
        data = [month_names[month-1], rent, utilities, food, stuff, other, total]
        for col, value in enumerate(data, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col > 1:  # Money columns
                cell.number_format = money_format
                cell.alignment = Alignment(horizontal='right')
    
    # Yearly total row
    total_row = 14
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 8):
        cell = ws_summary.cell(row=total_row, column=col)
        cell.value = f"=SUM({chr(64+col)}2:{chr(64+col)}13)"
        cell.number_format = money_format
        cell.font = Font(bold=True)
        cell.border = thin_border
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 12
    for col in 'BCDEFG':
        ws_summary.column_dimensions[col].width = 12
    
    # === Sheet 2: Per-Person Summary ===
    ws_person = wb.create_sheet("Per Person")
    
    headers = ['Person', 'Food', 'Utilities', 'Stuff', 'Other', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws_person.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for row, person in enumerate(config.USERS, 2):
        food = sum(FoodExpense.get_total_by_person_and_month(person, selected_year, m) 
                   for m in range(1, 13))
        utility = sum(UtilityExpense.get_total_by_person_and_month(person, selected_year, m) 
                      for m in range(1, 13))
        stuff = sum(StuffExpense.get_total_by_person_and_month(person, selected_year, m) 
                    for m in range(1, 13))
        other = sum(OtherExpense.get_total_by_person_and_month(person, selected_year, m) 
                    for m in range(1, 13))
        total = food + utility + stuff + other
        
        data = [person, food, utility, stuff, other, total]
        for col, value in enumerate(data, 1):
            cell = ws_person.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col > 1:
                cell.number_format = money_format
    
    # === Sheet 3: Food Expenses ===
    _create_expense_sheet(wb, "Food", FoodExpense.get_all(), date_format, money_format, thin_border, header_font, header_fill)
    
    # === Sheet 4: Utilities ===
    _create_utility_sheet(wb, UtilityExpense.get_all(), date_format, money_format, thin_border, header_font, header_fill)
    
    # === Sheet 5: Stuff ===
    _create_stuff_sheet(wb, StuffExpense.get_all(), date_format, money_format, thin_border, header_font, header_fill)
    
    # === Sheet 6: Other ===
    _create_expense_sheet(wb, "Other", OtherExpense.get_all(), date_format, money_format, thin_border, header_font, header_fill)
    
    # Save to bytes buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Generate filename
    filename = f"expenses_{selected_year}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def _create_expense_sheet(wb, name, expenses, date_format, money_format, border, header_font, header_fill):
    """Create a sheet for simple expenses (Food, Other)."""
    ws = wb.create_sheet(name)
    
    headers = ['Date', 'Name', 'Paid By', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=expense.expense_date).number_format = date_format
        ws.cell(row=row, column=2, value=expense.name)
        ws.cell(row=row, column=3, value=expense.paid_by)
        cell = ws.cell(row=row, column=4, value=expense.amount)
        cell.number_format = money_format
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12


def _create_utility_sheet(wb, expenses, date_format, money_format, border, header_font, header_fill):
    """Create sheet for utility expenses with type column."""
    ws = wb.create_sheet("Utilities")
    
    headers = ['Date', 'Type', 'Name', 'Paid By', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=expense.expense_date).number_format = date_format
        ws.cell(row=row, column=2, value=expense.utility_type)
        ws.cell(row=row, column=3, value=expense.name)
        ws.cell(row=row, column=4, value=expense.paid_by)
        cell = ws.cell(row=row, column=5, value=expense.amount)
        cell.number_format = money_format
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border


def _create_stuff_sheet(wb, expenses, date_format, money_format, border, header_font, header_fill):
    """Create sheet for stuff expenses with type column."""
    ws = wb.create_sheet("Stuff")
    
    headers = ['Date', 'Type', 'Name', 'Paid By', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row, expense in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=expense.expense_date).number_format = date_format
        ws.cell(row=row, column=2, value=expense.stuff_type)
        ws.cell(row=row, column=3, value=expense.name)
        ws.cell(row=row, column=4, value=expense.paid_by)
        cell = ws.cell(row=row, column=5, value=expense.amount)
        cell.number_format = money_format
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
